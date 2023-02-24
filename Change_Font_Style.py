import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row = 1, column = 1).value = "CodeMaster7000"
sheet.cell(row = 1, column = 1).font = Font(size = 24 )
sheet.cell(row = 2, column = 2).value = "CodeMaster7000"
sheet.cell(row = 2, column = 2).font = Font(size = 24, italic = True)
sheet.cell(row = 3, column = 3).value = "CodeMaster7000"
sheet.cell(row = 3, column = 3).font = Font(size = 24, bold = True)
sheet.cell(row = 4, column = 4).value = "CodeMaster7000"
sheet.cell(row = 4, column = 4).font = Font(size = 24, name = 'Times New Roman')
wb.save('sample.xlsx')
