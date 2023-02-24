import openpyxl 
wb = openpyxl.Workbook() 
sheet = wb.active 
sheet.merge_cells('A2:D3') 
sheet.cell(row = 1, column = 1).value = 'These cells have been merged.'
wb.save('sample.xlsx')
